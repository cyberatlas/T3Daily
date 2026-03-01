#!/usr/bin/env python3
"""
Parse 80/20 workout CSV or XLSX files into the project's JSON and CSV library format.

Supports both CSV and XLSX input. XLSX is preferred because it preserves
hyperlinks to .FIT workout files that are lost in CSV export.

Usage:
    pyenv activate T3Daily
    python scripts/parse_workout_csv.py <input_file> <sport> [--merge]

Arguments:
    input_file  Path to CSV or XLSX file (e.g., ~/Documents/8020runlibrary.xlsx)
    sport       One of: run, bike, swim
    --merge     Also merge results into data/plan/8020_workout_library.json

Examples:
    python scripts/parse_workout_csv.py ~/Documents/8020runlibrary.xlsx run --merge
    python scripts/parse_workout_csv.py ~/Documents/8020bikelibrary.csv bike --merge
    python scripts/parse_workout_csv.py ~/Documents/8020swimlibrary.xlsx swim --merge
"""

import csv
import json
import re
import sys
import os

# Workout type lookup: prefix -> (full name, intensity)
# Sorted by prefix length (longest first) so e.g. RLFF matches before RL
WORKOUT_TYPES = {
    # Run
    "RLFF": ("Running Long Fast Finish", "moderate"),
    "RLSP": ("Running Long Speed Play", "moderate"),
    "RLMS": ("Running Long Marathon Specific", "moderate"),
    "RRe":  ("Running Recovery", "low"),
    "RF":   ("Running Foundation", "low"),
    "RAe":  ("Running Aerobic", "low"),
    "RL":   ("Running Long", "low"),
    "RTa":  ("Running Taper", "low"),
    "RCI":  ("Running Cruise Intervals", "moderate"),
    "RT":   ("Running Tempo", "moderate"),
    "RFF":  ("Running Fast Finish", "moderate"),
    "RSP":  ("Running Speed Play", "moderate"),
    "RHR":  ("Running Hill Repeats", "high"),
    "RMI":  ("Running Mixed Intervals", "high"),
    "RAn":  ("Running Anaerobic", "high"),
    "RSI":  ("Running Speed Intervals", "high"),
    "RLI":  ("Running Long Intervals", "high"),

    # Bike
    "CFo":  ("Cycling Foundation (Long)", "low"),
    "CFF":  ("Cycling Fast Finish", "moderate"),
    "CRe":  ("Cycling Recovery", "low"),
    "CSP":  ("Cycling Speed Play", "moderate"),
    "CAe":  ("Cycling Aerobic", "low"),
    "CAn":  ("Cycling Anaerobic", "high"),
    "CCI":  ("Cycling Cruise Intervals", "moderate"),
    "CF":   ("Cycling Foundation", "low"),
    "CT":   ("Cycling Tempo", "moderate"),
    "CMI":  ("Cycling Mixed Intervals", "high"),
    "CSI":  ("Cycling Speed Intervals", "high"),
    "CLI":  ("Cycling Long Intervals", "high"),
    "CTa":  ("Cycling Taper", "low"),

    # Swim
    "STT":  ("Swim Time Trial", "moderate"),
    "SMI":  ("Swim Mixed Intervals", "high"),
    "SCI":  ("Swim Cruise Intervals", "moderate"),
    "SSP":  ("Swim Speed Play", "moderate"),
    "SSI":  ("Swim Speed Intervals", "high"),
    "SAe":  ("Swim Aerobic", "low"),
    "STa":  ("Swim Taper", "low"),
    "SF":   ("Swim Foundation", "low"),
    "ST":   ("Swim Tempo", "moderate"),
    "SRe":  ("Swim Recovery", "low"),
}


def get_workout_type(code):
    """Determine workout name and intensity from code prefix."""
    for prefix in sorted(WORKOUT_TYPES.keys(), key=len, reverse=True):
        if code.startswith(prefix):
            name_base, intensity = WORKOUT_TYPES[prefix]
            num = code[len(prefix):]
            return f"{name_base} {num}", intensity
    return code, "unknown"


def parse_duration(dur_str):
    """Parse duration string into (value, unit) tuple."""
    dur_str = dur_str.strip()
    # "XX min" format
    m = re.match(r'(\d+)\s*min', dur_str)
    if m:
        return int(m.group(1)), "minutes"
    # "XX mi" or "XXmi" format
    m = re.match(r'(\d+)\s*mi', dur_str)
    if m:
        return int(m.group(1)), "miles"
    # "XXXX yds" or "XXXX yards" format (swim)
    m = re.match(r'(\d+)\s*y(?:ds|ards)?', dur_str, re.IGNORECASE)
    if m:
        return int(m.group(1)), "yards"
    # "XX m" meters (swim)
    m = re.match(r'(\d+)\s*m$', dur_str)
    if m:
        return int(m.group(1)), "meters"
    return dur_str, "unknown"


def extract_max_zone(description):
    """Find the highest zone number mentioned in a workout description."""
    zones = re.findall(r'Zone\s+(\d+)', description)
    if zones:
        return max(int(z) for z in zones)
    return None


def _build_workout(code, duration_raw, description, sport, fit_url=None):
    """Build a workout dict from parsed fields."""
    description = re.sub(r'\s+', ' ', description).strip()

    name, intensity = get_workout_type(code)
    duration_val, duration_unit = parse_duration(duration_raw)
    max_zone = extract_max_zone(description)

    zones_str = f"Zone 1-{max_zone}" if max_zone else ""

    workout = {
        "name": name,
        "sport": sport,
        "intensity": intensity,
        "duration": duration_raw,
        "duration_minutes": duration_val if duration_unit == "minutes" else None,
        "zones": zones_str,
        "max_zone": max_zone,
        "description": description,
    }

    if fit_url:
        workout["fit_file_url"] = fit_url

    if duration_unit == "miles":
        workout["duration_miles"] = duration_val
    elif duration_unit == "yards":
        workout["duration_yards"] = duration_val
    elif duration_unit == "meters":
        workout["duration_meters"] = duration_val

    return workout


def parse_csv(input_path, sport):
    """Parse a workout CSV file and return a dict of workouts."""
    workouts = {}

    with open(input_path, 'r') as f:
        reader = csv.reader(f)
        header = next(reader)  # Skip header row
        print(f"Header: {header}")

        for row in reader:
            if len(row) < 3 or not row[0].strip():
                continue

            code = row[0].strip()
            duration_raw = row[1].strip()
            description = row[2].strip()

            workouts[code] = _build_workout(code, duration_raw, description, sport)

    return workouts


def parse_xlsx(input_path, sport):
    """Parse a workout XLSX file, extracting hyperlinks to .FIT files."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl is required for xlsx parsing.")
        print("Install with: pip install openpyxl")
        sys.exit(1)

    workouts = {}
    wb = load_workbook(input_path)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=False))
    header = [cell.value for cell in rows[0]]
    print(f"Header: {header}")

    links_found = 0
    for row in rows[1:]:
        code_cell = row[0]
        code = str(code_cell.value).strip() if code_cell.value else ""
        if not code:
            continue

        duration_raw = str(row[1].value).strip() if row[1].value else ""
        description = str(row[2].value).strip() if row[2].value else ""

        # Extract hyperlink from the code cell (column A)
        fit_url = None
        if code_cell.hyperlink:
            fit_url = code_cell.hyperlink.target
            links_found += 1

        workouts[code] = _build_workout(code, duration_raw, description, sport, fit_url)

    print(f"Extracted {links_found} .FIT file hyperlinks from xlsx")
    wb.close()
    return workouts


def save_outputs(workouts, sport, project_dir):
    """Save parsed workouts as JSON and CSV."""
    output_json = os.path.join(project_dir, f"data/plan/8020_{sport}_workouts.json")
    output_csv = os.path.join(project_dir, f"data/plan/8020_{sport}_workouts.csv")

    # JSON
    output = {
        "metadata": {
            "source": "80/20 Endurance",
            "source_url": "https://www.8020endurance.com/8020-workout-library/",
            "attribution": "Workout descriptions sourced from 80/20 Endurance. Used with attribution for non-commercial purposes.",
            "sport": sport,
            "version": "1.0",
            "total_workouts": len(workouts),
        },
        "workouts": workouts,
    }

    with open(output_json, 'w') as f:
        json.dump(output, f, indent=2)
    print(f"Saved {len(workouts)} workouts to {output_json}")

    # CSV
    with open(output_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([
            'code', 'name', 'sport', 'intensity', 'duration',
            'duration_minutes', 'max_zone', 'zones', 'fit_file_url', 'description'
        ])
        for code in sorted(workouts.keys()):
            w = workouts[code]
            writer.writerow([
                code, w['name'], w['sport'], w['intensity'], w['duration'],
                w.get('duration_minutes') or '', w.get('max_zone') or '',
                w['zones'], w.get('fit_file_url') or '', w['description']
            ])
    print(f"Saved {len(workouts)} workouts to {output_csv}")


def merge_into_library(workouts, sport, project_dir):
    """Merge parsed workouts into the main workout library skeleton."""
    lib_path = os.path.join(project_dir, "data/plan/8020_workout_library.json")

    if not os.path.exists(lib_path):
        print(f"Library not found at {lib_path} — skipping merge")
        return

    with open(lib_path) as f:
        library = json.load(f)

    updated = 0
    added = 0
    for code, workout in workouts.items():
        if code in library['workouts']:
            # Update existing entry
            lib_entry = library['workouts'][code]
            lib_entry['name'] = workout['name']
            lib_entry['intensity'] = workout['intensity']
            lib_entry['duration_minutes'] = workout.get('duration_minutes')
            lib_entry['zones'] = workout['zones']
            lib_entry['description'] = workout['description']
            if workout.get('fit_file_url'):
                lib_entry['fit_file_url'] = workout['fit_file_url']
            updated += 1
        else:
            # Add new entry (workout exists in full library but not in plan skeleton)
            entry = {
                "name": workout['name'],
                "sport": sport,
                "intensity": workout['intensity'],
                "duration_minutes": workout.get('duration_minutes'),
                "zones": workout['zones'],
                "description": workout['description'],
            }
            if workout.get('fit_file_url'):
                entry['fit_file_url'] = workout['fit_file_url']
            library['workouts'][code] = entry
            added += 1

    with open(lib_path, 'w') as f:
        json.dump(library, f, indent=2)

    print(f"Merged into {lib_path}: {updated} updated, {added} added")

    # Report remaining gaps
    empty = {s: 0 for s in ['swim', 'bike', 'run']}
    filled = {s: 0 for s in ['swim', 'bike', 'run']}
    for w in library['workouts'].values():
        s = w.get('sport', 'unknown')
        if s in empty:
            if w.get('name'):
                filled[s] += 1
            else:
                empty[s] += 1

    print(f"\nLibrary status:")
    for s in ['swim', 'bike', 'run']:
        status = "✅" if empty[s] == 0 else "⏳"
        print(f"  {status} {s}: {filled[s]} filled, {empty[s]} empty")


def print_summary(workouts):
    """Print a summary of parsed workout types."""
    types_seen = {}
    for code in sorted(workouts.keys()):
        for prefix in sorted(WORKOUT_TYPES.keys(), key=len, reverse=True):
            if code.startswith(prefix):
                types_seen.setdefault(prefix, []).append(code)
                break
        else:
            types_seen.setdefault("???", []).append(code)

    print(f"\nWorkout types found:")
    for prefix in sorted(types_seen.keys()):
        codes = types_seen[prefix]
        if prefix in WORKOUT_TYPES:
            name, intensity = WORKOUT_TYPES[prefix]
            print(f"  {prefix:6s} ({intensity:8s}) {name:35s} — {len(codes)} workouts")
        else:
            print(f"  {prefix:6s} (unknown ) {'???':35s} — {len(codes)} workouts: {codes}")


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    input_file = sys.argv[1]
    sport = sys.argv[2]
    do_merge = '--merge' in sys.argv

    if sport not in ('run', 'bike', 'swim'):
        print(f"Error: sport must be 'run', 'bike', or 'swim' (got '{sport}')")
        sys.exit(1)

    if not os.path.exists(input_file):
        print(f"Error: file not found: {input_file}")
        sys.exit(1)

    # Find project root (where this script lives in scripts/)
    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Auto-detect file type
    is_xlsx = input_file.lower().endswith('.xlsx')
    print(f"Parsing {input_file} as {sport} workouts ({'xlsx' if is_xlsx else 'csv'} format)...")

    if is_xlsx:
        workouts = parse_xlsx(input_file, sport)
    else:
        workouts = parse_csv(input_file, sport)
    print(f"Parsed {len(workouts)} {sport} workouts")

    print_summary(workouts)
    save_outputs(workouts, sport, project_dir)

    if do_merge:
        print("")
        merge_into_library(workouts, sport, project_dir)

    # Print a few examples
    print(f"\nExample workouts:")
    examples = list(workouts.keys())[:5]
    for code in examples:
        w = workouts[code]
        print(f"  {code:8s} | {w['intensity']:8s} | {w['duration']:8s} | {w['name']}")
        if w.get('fit_file_url'):
            print(f"           .FIT: {w['fit_file_url']}")
        print(f"           {w['description'][:80]}")


if __name__ == '__main__':
    main()
